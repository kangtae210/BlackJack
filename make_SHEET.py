import openpyxl

def make_sheet():
    
    locker = openpyxl.load_workbook(('locker.xlsx'))
    name = input("사물함 이름 =>")
    count = int(input("사물함 개수 =>"))
    locker.create_sheet(name)
    this_sheet = locker[name]
    this_sheet['A1'] = '사물함번호'
    this_sheet['B1'] = '사용자1_이름'
    this_sheet['C1'] = '사용자1_학번'
    this_sheet['D1'] = '사용자1_휴대폰번호'
    this_sheet['E1'] = '사용자2_이름'
    this_sheet['F1'] = '사용자2_학번'
    this_sheet['G1'] = '사용자2_휴대폰번호'
    this_sheet['H1'] = '상태'    
    this_sheet['I1'] = '비고'      
    for i in range(count):
        this_sheet.cell(row=i+2, column= 1).value = str(i+1)
    locker.save('locker.xlsx')
    locker.close()


        

